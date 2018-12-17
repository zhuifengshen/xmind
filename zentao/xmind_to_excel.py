#!/usr/bin/env python
# _*_ coding:utf-8 _*_

import os
import xmind
import logging
import openpyxl
import xlrd
from xlutils.copy import copy as xlscopy


class XMindToExcel(object):

    def __init__(self, xmind, excel, template):
        self.source_xmind_file = xmind
        self.templet_excel_file = template
        self.target_excel_file = excel

    def parse_xmind(self, xmind_file):
        workbook = xmind.load(xmind_file)
        primary_sheet = workbook.getPrimarySheet()  # get the first sheet
        root_topic = primary_sheet.getRootTopic()  # get the root topic of this sheet
        case_group_topics = root_topic.getSubTopics()

        case = dict(
            group="",     # 用例分组名
            title="",     # 用例标题
            priority="",  # 用例优先级
            prestep="",    # 前置条件
            step="",      # 用例步骤
            expect="",    # 期望结果
            remarks="",   # 备注
            upload="是")   # 上传

        for case_group_topic in case_group_topics:
            case_group_name = case_group_topic.getTitle()  # case group
            case["group"] = case_group_name
            case_title_topics = case_group_topic.getSubTopics()
            if not case_title_topics:
                break
            for case_title_topic in case_title_topics:
                case_title_name = case_title_topic.getTitle()  # case title
                case["title"] = case_title_name
                count_case_num = 1
                case_mainstep_topics = case_title_topic.getSubTopics()
                if not case_mainstep_topics:
                    break
                for case_mainstep_topic in case_mainstep_topics:
                    case_mainstep_name = case_mainstep_topic.getTitle()  # case mainstep
                    case_prestep_name = ""
                    if case_mainstep_topic.getNotes():
                        case_prestep_name = case_mainstep_topic.getNotes()  # case prestep
                    case["prestep"] = case_prestep_name
                    case_substep_topics = case_mainstep_topic.getSubTopics()
                    if not case_substep_topics:
                        break
                    for case_substep_topic in case_substep_topics:
                        case_substep_name = case_substep_topic.getTitle()  # case substep
                        case_priority_name = "中"
                        if case_substep_topic.getMarkers():
                            case_priority_name = case_substep_topic.getMarkers()[0].getMarkerId()  # case priority
                            case_priority_name = self.switch_priority(str(case_priority_name))
                        case["priority"] = case_priority_name
                        case["step"] = self.case_step(case_mainstep_name, case_substep_name)
                        case_expect_topics = case_substep_topic.getSubTopics()
                        if not case_expect_topics:
                            break
                        for case_expect_topic in case_expect_topics:
                            case_expect_name = case_expect_topic.getTitle()  # case expect
                            case["expect"] = case_expect_name
                            case_remarks_name = ""
                            if case_expect_topic.getNotes():
                                case_remarks_name = case_expect_topic.getNotes()
                            case["remarks"] = case_remarks_name
                            case_upload_topics = case_expect_topic.getSubTopics()
                            if case_upload_topics and case_upload_topics[0].getTitle() == "不上传":  # case upload
                                case["upload"] = "否"
                            else:
                                case["upload"] = "是"
                            entity_case = case.copy()
                            if count_case_num > 1:
                                entity_case["title"] = self.case_title(entity_case["title"], count_case_num)
                            count_case_num = count_case_num + 1
                            yield entity_case

    def main(self):
        if not self.source_xmind_file or not self.templet_excel_file or not self.target_excel_file:
            return False
        if os.path.splitext(self.templet_excel_file)[1] == '.xlsx':
            logging.info('XMind测试用例文件转化为Excel xlsx格式文件...')
            workbook = openpyxl.load_workbook(filename=self.templet_excel_file)
            sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])
            last_row_num = sheet.max_row
        else:
            logging.info('XMind测试用例文件转化为Excel xls格式文件...')
            template_workbook = xlrd.open_workbook(self.templet_excel_file, encoding_override='GB2312', formatting_info=True)
            template_sheet = template_workbook.sheet_by_index(0)
            last_row_num = template_sheet.nrows
            workbook = xlscopy(template_workbook)
            sheet = workbook.get_sheet(0)

        for case in self.parse_xmind(self.source_xmind_file):
            logging.debug("TestCase\n功能模块：%(group)s, 用例标题：%(title)s, 优先级：%(priority)s, 前置条件：%(prestep)s, 测试步骤：%(step)s, 预期结果：%(expect)s, 备注：%(remarks)s, 是否上传：%(upload)s" % case)
            last_row_num = self.add_case_to_workbook(case, sheet, last_row_num)
        workbook.save(self.target_excel_file)
        return True

    def add_case_to_workbook(self, case, sheet, row_num):
        """
        用例部分从第二行开始, 每行八列：功能模块、用例标题、优先级（默认中）、前置条件、执行步骤、预期结果、备注说明、是否上传（默认上传）
        """
        if os.path.splitext(self.templet_excel_file)[1] == '.xlsx':
            sheet.cell(row=row_num+1, column=1, value=case["group"])
            sheet.cell(row=row_num+1, column=2, value=case["title"])
            sheet.cell(row=row_num+1, column=3, value=case["priority"])
            sheet.cell(row=row_num+1, column=4, value=case["prestep"])
            sheet.cell(row=row_num+1, column=5, value=case["step"])
            sheet.cell(row=row_num+1, column=6, value=case["expect"])
            sheet.cell(row=row_num+1, column=7, value=case["remarks"])
            sheet.cell(row=row_num+1, column=8, value=case["upload"])
        else:
            sheet.write(row_num, 0, case["group"])
            sheet.write(row_num, 1, case["title"])
            sheet.write(row_num, 2, case["priority"])
            sheet.write(row_num, 3, case["prestep"])
            sheet.write(row_num, 4, case["step"])
            sheet.write(row_num, 5, case["expect"])
            sheet.write(row_num, 6, case["remarks"])
            sheet.write(row_num, 7, case["upload"])
        return row_num + 1

    @staticmethod
    def switch_priority(xmind_priority):
        if xmind_priority == "priority-1":
            return "高"
        elif xmind_priority == "priority-2":
            return "中"
        elif xmind_priority == "priority-3":
            return "低"
        else:
            return "中"

    @staticmethod
    def case_step(mainstep, substep):
        separator = "\n" + "=" * 10 + "\n"
        full_step = mainstep + separator + substep
        return full_step

    @staticmethod
    def case_title(title, num):
        return title + " 组用例-%d" % num


if __name__ == "__main__":
    xmind_Text_content = \
        "/Users/zhangchuzhao/Project/python/tmp/xmind-sdk-python/zentao/template/case_example.xmind"
    excel_file_Text_content = \
        "/Users/zhangchuzhao/Project/python/tmp/xmind-sdk-python/zentao/template/result.xlsx"
    templet_Text_content = \
        "/Users/zhangchuzhao/Project/python/tmp/xmind-sdk-python/zentao/template/case_template.xlsx"

    xmindParse = XMindToExcel(xmind_Text_content, excel_file_Text_content, templet_Text_content)
    xmindParse.main()
