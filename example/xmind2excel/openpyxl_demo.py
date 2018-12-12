#!/usr/bin/env python
# _*_ coding:utf-8 _*_

from openpyxl import Workbook
from datetime import datetime

"""
openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值
"""

wb = Workbook()

ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
ws['A2'] = datetime.now()

wb.save('sample.xlsx')
